from openpyxl import load_workbook
class DoExcel:
    def __init__(self,file_name,sheet):
        self.file_name=file_name
        self.sheet=sheet

    def read_excel(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet]
        test_data=[]
        for i in range(2,sheet.max_row+1):
            list_data=[]
            for j in range(1,8):
                list_data.append(sheet.cell(i,j).value)

            test_data.append(list_data)
        return test_data
    def write_excel(self,row,actual,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet]
        sheet.cell(row, 8).value = actual
        sheet.cell(row, 9).value = result
        wb.save(self.file_name)



if __name__ == '__main__':
    #实例化
    t=DoExcel("D:\BaiduNetdiskDownload\\aotu_interface_3\\test_data\\test.xlsx","Sheet1").read_excel()

    print(t)







